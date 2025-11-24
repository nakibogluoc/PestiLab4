from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Query
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any, Tuple
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
from openpyxl import load_workbook, Workbook
import io
import qrcode
import barcode
from barcode.writer import ImageWriter
import pytz
from io import BytesIO
import base64
import re

# ==== INIT ====
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

# MongoDB
MONGO_URL = os.getenv("MONGO_URL", "")
DB_NAME = os.getenv("DB_NAME", "pestilab")
client: Optional[AsyncIOMotorClient] = AsyncIOMotorClient(MONGO_URL) if MONGO_URL else None
db = client[DB_NAME] if client else None

# JWT
SECRET_KEY = os.getenv("SECRET_KEY", "laboratory-secret-key-2025")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 480

# TZ
ISTANBUL_TZ = pytz.timezone("Europe/Istanbul")

# FastAPI app + router
app = FastAPI()
api_router = APIRouter(prefix="/api")

@api_router.get("/health")
async def api_health_check():
    return {"ok": True, "service": "pestilab-api", "path": "/api/health"}

security = HTTPBearer()

# ==== MODELS ====
class UserRole(str):
    ADMIN = "admin"
    MANAGER = "manager"
    ANALYST = "analyst"
    READONLY = "readonly"

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    email: str
    role: str = "analyst"
    created_at: str = Field(default_factory=lambda: datetime.now(ISTANBUL_TZ).isoformat())

class UserCreate(BaseModel):
    username: str
    email: str
    password: str
    role: str = "analyst"

class UserLogin(BaseModel):
    username: str
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str
    user: User

class Compound(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    cas_number: str
    solvent: str
    stock_value: float
    stock_unit: str = "mg"
    critical_value: float = 100.0
    critical_unit: str = "mg"
    last_serial: int = 0
    notes: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(ISTANBUL_TZ).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(ISTANBUL_TZ).isoformat())

class CompoundCreate(BaseModel):
    name: str
    cas_number: str
    solvent: str
    stock_value: float
    stock_unit: str = "mg"
    critical_value: float = 100.0
    critical_unit: str = "mg"
    notes: Optional[str] = None

class CompoundUpdate(BaseModel):
    name: Optional[str] = None
    cas_number: Optional[str] = None
    solvent: Optional[str] = None
    stock_value: Optional[float] = None
    stock_unit: Optional[str] = None
    critical_value: Optional[float] = None
    critical_unit: Optional[str] = None
    notes: Optional[str] = None

class SolventDensity(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    solvent_name: str
    temperature_c: float
    density_g_per_ml: float
    created_at: str = Field(default_factory=lambda: datetime.now(ISTANBUL_TZ).isoformat())

class SolventDensityCreate(BaseModel):
    solvent_name: str
    temperature_c: float
    density_g_per_ml: float

class WeighingInput(BaseModel):
    compound_id: str
    weighed_amount: float
    purity: float = 100.0
    target_concentration: float
    concentration_mode: str = "mg/L"
    temperature_c: float = 25.0
    solvent: Optional[str] = None
    prepared_by: str
    mix_code: Optional[str] = None
    mix_code_show: bool = True
    label_code: Optional[str] = None
    label_code_source: str = "auto"

class Usage(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    compound_id: str
    compound_name: str
    cas_number: str
    weighed_amount: float
    purity: float
    actual_mass: float
    target_concentration: float
    concentration_mode: str
    required_volume: float
    required_solvent_mass: float
    actual_concentration: float
    deviation: float
    solvent: str
    temperature_c: float
    solvent_density: float
    remaining_stock: float
    remaining_stock_unit: str
    prepared_by: str
    mix_code: Optional[str] = None
    mix_code_show: bool = True
    label_code_used: Optional[str] = None
    label_code_source: str = "auto"
    created_at: str = Field(default_factory=lambda: datetime.now(ISTANBUL_TZ).isoformat())

class Label(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    compound_id: str
    usage_id: str
    label_code: str
    compound_name: str
    cas_number: str
    concentration: str
    prepared_by: str
    date: str
    qr_data: str
    created_at: str = Field(default_factory=lambda: datetime.now(ISTANBUL_TZ).isoformat())

class ExcelImportPreview(BaseModel):
    to_insert: List[Dict[str, Any]]
    to_update: List[Dict[str, Any]]
    to_skip: List[Dict[str, Any]]
    total_rows: int

class ExcelImportResponse(BaseModel):
    message: str
    compounds_added: int
    compounds_updated: int
    compounds_skipped: int
    densities_added: int = 0

# ==== HELPERS ====
def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    token = credentials.credentials
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if not username:
            raise HTTPException(status_code=401, detail="Invalid token")
        if db is None:
            raise HTTPException(status_code=500, detail="DB not configured")
        user = await db.users.find_one({"username": username}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return User(**user)
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Invalid token")

def normalize_string(s: str) -> str:
    if not s:
        return ""
    return " ".join(str(s).strip().upper().split())

def normalize_for_search(text: str) -> str:
    if not text:
        return ""
    char_map = {
        'İ': 'i','I': 'i','ı': 'i','Ğ': 'g','ğ': 'g','Ş': 's','ş': 's',
        'Ç': 'c','ç': 'c','Ö': 'o','ö': 'o','Ü': 'u','ü': 'u'
    }
    text = str(text).lower()
    for k, v in char_map.items():
        text = text.replace(k.lower(), v)
    return re.sub(r'[\s\-\(\),]', '', text)

def calculate_search_score(query: str, compound_name: str, cas_number: str) -> int:
    score = 0
    query_norm = normalize_for_search(query)
    name_norm = normalize_for_search(compound_name)
    cas_norm = normalize_for_search(cas_number)

    query_lower = query.lower()
    name_lower = compound_name.lower()
    cas_lower = cas_number.lower()

    if query_norm == name_norm or query_norm == cas_norm:
        score += 100
    if query_lower == name_lower or query_lower == cas_lower:
        score += 95
    if name_norm.startswith(query_norm):
        score += 60
    if cas_norm.startswith(query_norm):
        score += 60
    for word in name_lower.split():
        if word.startswith(query_lower):
            score += 50
    if query_norm in name_norm:
        score += 40
    if query_norm in cas_norm:
        score += 40
    if query_lower in name_lower:
        score += 35
    if query_lower in cas_lower:
        score += 35
    if len(query_norm) >= 2:
        score += min(len(query_norm) * 2, 20)
    if len(query_norm) < 4 and len(name_norm) > 20:
        score -= 5
    return score

def normalize_compound_name(name: str) -> str:
    char_map = {'İ':'I','ı':'i','Ğ':'G','ğ':'g','Ş':'S','ş':'s','Ç':'C','ç':'c','Ö':'O','ö':'o','Ü':'U','ü':'u'}
    normalized = ''
    for ch in name:
        if ch in char_map:
            normalized += char_map[ch]
        elif ch.isalpha():
            normalized += ch
    prefix = normalized[:3].upper()
    if len(prefix) < 3:
        prefix = prefix.ljust(3, 'X')
    return prefix

def find_column_by_aliases(headers: Dict[str, int], aliases: List[str]) -> Optional[int]:
    for header_name, col_idx in headers.items():
        normalized_header = normalize_string(header_name)
        for alias in aliases:
            if normalize_string(alias) == normalized_header:
                return col_idx
    return None

def interpolate_density(temperature: float, density_data: List[Dict[str, float]]) -> Tuple[float, bool]:
    if not density_data:
        return 0.8, False
    sorted_data = sorted(density_data, key=lambda x: x['temperature_c'])
    for d in sorted_data:
        if d['temperature_c'] == temperature:
            return d['density_g_per_ml'], False
    for i in range(len(sorted_data) - 1):
        t1, d1 = sorted_data[i]['temperature_c'], sorted_data[i]['density_g_per_ml']
        t2, d2 = sorted_data[i+1]['temperature_c'], sorted_data[i+1]['density_g_per_ml']
        if t1 <= temperature <= t2:
            density = d1 + (d2 - d1) * (temperature - t1) / (t2 - t1)
            return density, False
    if temperature < sorted_data[0]['temperature_c']:
        t1, d1 = sorted_data[0]['temperature_c'], sorted_data[0]['density_g_per_ml']
        t2, d2 = sorted_data[1]['temperature_c'], sorted_data[1]['density_g_per_ml']
    else:
        t1, d1 = sorted_data[-2]['temperature_c'], sorted_data[-2]['density_g_per_ml']
        t2, d2 = sorted_data[-1]['temperature_c'], sorted_data[-1]['density_g_per_ml']
    density = d1 + (d2 - d1) * (temperature - t1) / (t2 - t1)
    return density, True

def generate_qr_code(data: str) -> str:
    qr = qrcode.QRCode(version=1, box_size=10, border=1)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    buffer.seek(0)
    return base64.b64encode(buffer.getvalue()).decode()

def generate_barcode(code: str) -> str:
    buffer = BytesIO()
    code128 = barcode.get("code128", code, writer=ImageWriter())
    code128.write(buffer, {"write_text": False, "module_height": 8, "module_width": 0.2})
    buffer.seek(0)
    return base64.b64encode(buffer.getvalue()).decode()

# ==== AUTH ====
@api_router.post("/auth/register", response_model=User)
async def register(user_data: UserCreate, current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admin can create users")
    if db is None:
        raise HTTPException(status_code=500, detail="DB not configured")
    existing_user = await db.users.find_one({"username": user_data.username})
    if existing_user:
        raise HTTPException(status_code=400, detail="Username already exists")
    hashed_password = bcrypt.hashpw(user_data.password.encode("utf-8"), bcrypt.gensalt())
    user = User(username=user_data.username, email=user_data.email, role=user_data.role)
    doc = user.model_dump()
    doc["password"] = hashed_password.decode("utf-8")
    await db.users.insert_one(doc)
    return user

@api_router.post("/auth/login", response_model=Token)
async def login(login_data: UserLogin):
    if db is None:
        raise HTTPException(status_code=500, detail="DB not configured")
    user = await db.users.find_one({"username": login_data.username})
    if not user:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    if not bcrypt.checkpw(login_data.password.encode("utf-8"), user["password"].encode("utf-8")):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    access_token = create_access_token(data={"sub": user["username"]})
    user_obj = User(**{k: v for k, v in user.items() if k != "password"})
    return Token(access_token=access_token, token_type="bearer", user=user_obj)

@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: User = Depends(get_current_user)):
    return current_user

@api_router.get("/users", response_model=List[User])
async def get_users(current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "manager"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    if db is None:
        raise HTTPException(status_code=500, detail="DB not configured")
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    return [User(**u) for u in users]

# ==== SOLVENT DENSITY ====
@api_router.post("/solvent-densities", response_model=SolventDensity)
async def create_solvent_density(data: SolventDensityCreate, current_user: User = Depends(get_current_user)):
    if current_user.role == "readonly":
        raise HTTPException(status_code=403, detail="Read-only users cannot create density data")
    if db is None:
        raise HTTPException(status_code=500, detail="DB not configured")
    density = SolventDensity(**data.model_dump())
    await db.solvent_densities.insert_one(density.model_dump())
    return density

@api_router.get("/solvent-densities", response_model=List[SolventDensity])
async def get_solvent_densities(current_user: User = Depends(get_current_user)):
    if db is None:
        raise HTTPException(status_code=500, detail="DB not configured")
    densities = await db.solvent_densities.find({}, {"_id": 0}).to_list(1000)
    return [SolventDensity(**d) for d in densities]

@api_router.get("/solvent-densities/{solvent_name}/at/{temperature}")
async def get_density_at_temperature(solvent_name: str, temperature: float, current_user: User = Depends(get_current_user)):
    if db is None:
        raise HTTPException(status_code=500, detail="DB not configured")
    density_data = await db.solvent_densities.find({"solvent_name": solvent_name}, {"_id": 0}).to_list(100)
    if not density_data:
        raise HTTPException(status_code=404, detail=f"No density data found for solvent: {solvent_name}")
    density, is_extrapolated = interpolate_density(temperature, density_data)
    return {
        "solvent_name": solvent_name,
        "temperature_c": temperature,
        "density_g_per_ml": round(density, 4),
        "is_extrapolated": is_extrapolated,
        "warning": "Extrapolated density - outside measured range" if is_extrapolated else None
    }

# ==== COMPOUNDS ====
@api_router.post("/compounds", response_model=Compound)
async def create_compound(compound_data: CompoundCreate, current_user: User = Depends(get_current_user)):
    if current_user.role == "readonly":
        raise HTTPException(status_code=403, detail="Read-only users cannot create compounds")
    if db is None:
        raise HTTPException(status_code=500, detail="DB not configured")
    compound = Compound(**compound_data.model_dump())
    await db.compounds.insert_one(compound.model_dump())
    await db.audit_logs.insert_one({
        "id": str(uuid.uuid4()),
        "user": current_user.username,
        "action": "create_compound",
        "compound_id": compound.id,
        "compound_name": compound.name,
        "timestamp": datetime.now(ISTANBUL_TZ).isoformat()
    })
    return compound

@api_router.get("/compounds", response_model=List[Compound])
async def get_compounds(current_user: User = Depends(get_current_user)):
    if db is None:
        raise HTTPException(status_code=500, detail="DB not configured")
    compounds = await db.compounds.find({}, {"_id": 0}).to_list(10000)
    return [Compound(**c) for c in compounds]

@api_router.get("/compounds/{compound_id}", response_model=Compound)
async def get_compound(compound_id: str, current_user: User = Depends(get_current_user)):
    if db is None:
        raise HTTPException(status_code=500, detail="DB not configured")
    compound = await db.compounds.find_one({"id": compound_id}, {"_id": 0})
    if not compound:
        raise HTTPException(status_code=404, detail="Compound not found")
    return Compound(**compound)

@api_router.put("/compounds/{compound_id}", response_model=Compound)
async def update_compound(compound_id: str, update_data: CompoundUpdate, current_user: User = Depends(get_current_user)):
    if current_user.role == "readonly":
        raise HTTPException(status_code=403, detail="Read-only users cannot update compounds")
    if db is None:
        raise HTTPException(status_code=500, detail="DB not configured")
    compound = await db.compounds.find_one({"id": compound_id}, {"_id": 0})
    if not compound:
        raise HTTPException(status_code=404, detail="Compound not found")
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    update_dict["updated_at"] = datetime.now(ISTANBUL_TZ).isoformat()
    await db.compounds.update_one({"id": compound_id}, {"$set": update_dict})
    await db.audit_logs.insert_one({
        "id": str(uuid.uuid4()),
        "user": current_user.username,
        "action": "update_compound",
        "compound_id": compound_id,
        "changes": update_dict,
        "timestamp": datetime.now(ISTANBUL_TZ).isoformat()
    })
    updated_compound = await db.compounds.find_one({"id": compound_id}, {"_id": 0})
    return Compound(**updated_compound)

@api_router.delete("/compounds/{compound_id}")
async def delete_compound(compound_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Only admin can delete compounds")
    if db is None:
        raise HTTPException(status_code=500, detail="DB not configured")
    result = await db.compounds.delete_one({"id": compound_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Compound not found")
    await db.audit_logs.insert_one({
        "id": str(uuid.uuid4()),
        "user": current_user.username,
        "action": "delete_compound",
        "compound_id": compound_id,
        "timestamp": datetime.now(ISTANBUL_TZ).isoformat()
    })
    return {"message": "Compound deleted successfully"}

# ==== EXCEL IMPORT ====
@api_router.post("/compounds/import/preview", response_model=ExcelImportPreview)
async def preview_excel_import(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    if current_user.role == "readonly":
        raise HTTPException(status_code=403, detail="Read-only users cannot import data")
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only Excel files are supported")
    if db is None:
        raise HTTPException(status_code=500, detail="DB not configured")

    contents = await file.read()
    workbook = load_workbook(filename=io.BytesIO(contents), read_only=True)

    to_insert, to_update, to_skip = [], [], []

    sheet = None
    for sheet_name in workbook.sheetnames:
        if "compound" in sheet_name.lower() or sheet_name == workbook.sheetnames[0]:
            sheet = workbook[sheet_name]
            break
    if not sheet:
        sheet = workbook.active

    name_aliases = ["Analit Adı", "Compound", "Compound Name", "Name"]
    cas_aliases = ["CAS", "CAS No", "CAS Number"]
    solvent_aliases = ["Solvent", "Çözücü", "Önerilen Solvent", "Default Solvent"]

    header_row, headers = None, {}
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=100), start=1):
        row_values = [cell.value for cell in row if cell.value]
        if len(row_values) >= 2:
            for cell in row:
                if cell.value:
                    val_str = str(cell.value)
                    if any(alias in val_str for alias in name_aliases + cas_aliases):
                        header_row = row_idx
                        for c in row:
                            if c.value:
                                headers[c.value] = c.column
                        break
        if header_row:
            break
    if not header_row:
        raise HTTPException(status_code=400, detail="Could not find header row with required columns")

    name_col = find_column_by_aliases(headers, name_aliases)
    cas_col = find_column_by_aliases(headers, cas_aliases)
    solvent_col = find_column_by_aliases(headers, solvent_aliases)
    if not name_col or not cas_col:
        raise HTTPException(status_code=400, detail=f"Required columns not found. Headers found: {list(headers.keys())}")

    for row in sheet.iter_rows(min_row=header_row + 1, max_row=header_row + 500):
        name = row[name_col - 1].value if name_col else None
        cas = row[cas_col - 1].value if cas_col else None
        solvent = row[solvent_col - 1].value if solvent_col else "Acetone"
        if not name or not cas or str(cas).lower() == "nan" or str(name).startswith("="):
            continue
        name = str(name).strip()
        cas = str(cas).strip().upper()
        solvent = str(solvent).strip() if solvent else "Acetone"
        existing = await db.compounds.find_one({"cas_number": cas})
        compound_data = {
            "name": name, "cas_number": cas, "solvent": solvent,
            "stock_value": 1000.0, "stock_unit": "mg", "critical_value": 100.0
        }
        if existing:
            compound_data["id"] = existing["id"]
            to_update.append(compound_data)
        else:
            to_insert.append(compound_data)

    return ExcelImportPreview(
        to_insert=to_insert[:50],
        to_update=to_update[:50],
        to_skip=to_skip[:50],
        total_rows=len(to_insert) + len(to_update) + len(to_skip)
    )

@api_router.post("/compounds/import", response_model=ExcelImportResponse)
async def import_compounds(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    if current_user.role == "readonly":
        raise HTTPException(status_code=403, detail="Read-only users cannot import data")
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only Excel files are supported")
    if db is None:
        raise HTTPException(status_code=500, detail="DB not configured")

    contents = await file.read()
    workbook = load_workbook(filename=io.BytesIO(contents), read_only=True)

    added = updated = skipped = densities_added = 0

    sheet = None
    for sheet_name in workbook.sheetnames:
        if "compound" in sheet_name.lower() or sheet_name == workbook.sheetnames[0]:
            sheet = workbook[sheet_name]
            break
    if not sheet:
        sheet = workbook.active

    name_aliases = ["Analit Adı", "Compound", "Compound Name", "Name"]
    cas_aliases = ["CAS", "CAS No", "CAS Number"]
    solvent_aliases = ["Sol
